#!/usr/bin/env python3
# -*- coding: utf-8 -*-

'''
利用openpyxl操作excel
文档地址: http://openpyxl.readthedocs.io/en/default/#
tutorial地址: http://www.pythonexcel.com/openpyxl.php
测试后发现,不支持openoffice格式(.ods),只支持 .xlsx
'''

import openpyxl
import openpyxl.utils
from openpyxl.compat import range

src_excel_path = './py_src.xlsx'

# 加载已有的工作簿
wb = openpyxl.load_workbook(src_excel_path)
print(type(wb))

# 创建新工作簿
# newWb = openpyxl.Workbook()
# newWb.create_sheet("hello openpyxl", 0)
# newWb.save("py_src_new.xlsx")

try:
    # 创建工作表
    newSheet = wb.create_sheet('工作表4')
    wb.save(src_excel_path)

    sheetNames = wb.get_sheet_names()  # 获取所有的工作表名
    sheetCount = len(sheetNames)
    print("总共有 %s 张表" % sheetCount)

    # 删除工作表
    # 批量删除注意从最后一张表开始删除
    if sheetCount > 1:
        for index in range(sheetCount - 1, 0, -1):
            print("正在删除表 %s" % sheetNames[index])
            wb.remove_sheet(wb[sheetNames[index]])
            # wb.remove(wb.get_sheet_by_name(sheetNames[index]))
        print("删除操作过后,总共有 %s 张表" % len(wb.get_sheet_names()))

    # 获取某个工作表
    # sheet1 = wb.get_sheet_by_name('工作表1')
    sheet1 = wb['工作表1']

    # 获取有数据的区域最大行号列号: sum(1 for _ in sheet1.columns)  sum(1 for _ in sheet1.rows)
    print("sheet1 相关信息 ", type(sheet1), sheet1.title, sum(1 for _ in sheet1.columns), sum(1 for _ in sheet1.rows))
    print("max_column", sheet1.max_column, sheet1.max_row, len(sheet1['A']))
    # colA = tuple(sheet1.columns)[:1]
    # print("col length: ", len(colA))

    # 获取指定单元格对象
    cell = sheet1['a1']
    print("cell ", cell.column, cell.row, " value is : ", cell.value)
    # 方式2
    cell2 = sheet1.cell(row=1, column=2)  # 获取B1单元格
    print("cell B1 value is : %s " % (cell2.value))

    # 写入值,记得要save()保存
    # 方式1: 通过value属性
    cell.value = 'hello'
    wb.save(src_excel_path)

    # 方式2: 直接给cell赋值
    sheet1['a1'] = "world"
    wb.save(src_excel_path)

    # 切片,获取指定单元格组成的区域
    area = tuple(sheet1['a1':'c3'])
    print("length = ", len(area))

    # 选择第1到9行的区域
    for x in range(1, 10):
        print(x, sheet1.cell(row=x, column=1).value)

    # 从当前区域的下一行开始,插入10行,每行添加0~9共计10个单元格数据
    for row in range(10):
        sheet1.append(range(10))

    # 在当前有效区域的下一行追加一行数据,需要传入tuple,list 或 generator , 从A1单元格开始填充
    sheet1.append(["append operation", "by openpyxl"])
    wb.save(src_excel_path)

except KeyError as e:
    print("error ", e)
# except ValueError as e:
#     print("ValueError ", e)

# 列字母和数字之间的转换
print(openpyxl.utils.get_column_letter(1), openpyxl.utils.column_index_from_string('aa'))
