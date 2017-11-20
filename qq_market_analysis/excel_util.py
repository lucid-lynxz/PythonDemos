#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
import openpyxl.utils
from openpyxl import Workbook
from openpyxl.compat import range
import time

wb = ""
columnSize = 0
targetFileName = 'qq_market.xlsx'


def getWorkBook(fileName=targetFileName):
    global wb, targetFileName
    targetFileName = fileName
    try:
        if not wb:
            wb = openpyxl.load_workbook(targetFileName)
    except FileNotFoundError as e:
        print("FileNotFoundError ", e)
        wb = Workbook(fileName)
        wb.save(fileName)
        wb = openpyxl.load_workbook(fileName)


# 初始化excel操作, 写入列名称
def initWorkBook(*columnName, fileName=targetFileName):
    print("excel_util initWorkBook")
    global wb, columnSize
    getWorkBook(fileName)

    # 删除工作表,从最后一张表开始删除
    sheetNames = wb.get_sheet_names()  # 获取所有的工作表名
    sheetCount = len(sheetNames)
    print("总共有 %s 张表" % sheetCount)
    if sheetCount >= 1:
        for index in range(sheetCount - 1, -1, -1):
            print("正在删除表 %s" % sheetNames[index])
            wb.remove_sheet(wb[sheetNames[index]])
            # wb.remove(wb.get_sheet_by_name(sheetNames[index]))
        print("删除操作过后,总共有 %s 张表" % len(wb.get_sheet_names()))

    # 创建新表并写入列名
    columnSize = len(columnName)
    sheet = wb.create_sheet("market_%s" % (time.strftime("%Y%m%d%H%M%s", time.localtime(time.time()))))
    print("创建表名:%s  %s" % (sheet.title, columnSize))
    if columnSize > 0:
        try:
            for x in range(0, columnSize):
                columnIndex = x + 1
                # print("currentIndex=", columnIndex, columnSize, columnName[x])
                sheet.cell(row=1, column=columnIndex).value = "%s" % columnName[x]
        except Exception as e:
            print("error", e)
    wb.save(fileName)
    wb.close()


# 在表格的下一行追加一行数据
def appendItem(*info):
    global wb, targetFileName
    getWorkBook(targetFileName)
    sheetNames = wb.get_sheet_names()  # 获取所有的工作表名
    sheetCount = len(sheetNames)
    if sheetCount >= 1:
        sheet = wb[sheetNames[sheetCount - 1]]
        sheet.append(info)
        wb.save(targetFileName)
    else:
        wb.close()


# 在excel指定行号数据中, 添加特定columnName 值(列名在首行已确定,若无,则在本行下一个空白单元格内填入)
def appendInfo(sheetName, row, value, columnName='localApkFileName'):
    global wb, targetFileName
    getWorkBook(targetFileName)
    if not sheetName:
        sheetNames = wb.get_sheet_names()
        sheetSize = len(sheetNames)
        sheetName = sheetNames[sheetSize - 1]
    sheet = wb.get_sheet_by_name(sheetName)
    if sheet:
        columnIndex = getColumnIndexByName(sheet, columnName)
        # columnIndex = 10
        print("ori value = ", sheet.cell(row=row, column=columnIndex).value)
        sheet.cell(row=row, column=columnIndex).value = value
        wb.save(targetFileName)
        print(columnIndex, value, targetFileName, "currentValue is ", sheet.cell(row=row, column=columnIndex).value)
        close()
    else:
        print("指定的sheet(%s)不存在,appendInfo失败,请重试" % sheetName)


columnIndexDict = {}


# 根据指定的列名称获取列号
# 若指定的列名不存在,则返回下一个空白单元格序号
# @param sheet openpyxl的Worksheet对象
def getColumnIndexByName(sheet, columnName='localApkFileName'):
    global columnIndexDict
    tempIndex = columnIndexDict.get(columnName, -1)
    if tempIndex > 0:
        return tempIndex

    columnIndex = -1
    if sheet:
        row1ColumnLen = len(sheet[1])

        for x in range(1, row1ColumnLen + 1):
            if sheet.cell(row=1, column=x).value == columnName:
                columnIndex = x
                columnIndexDict[columnName] = x
                break

        if columnIndex < 0:
            columnIndex = len(sheet[1]) + 1

        return columnIndex
    else:
        print("表格不存在,请重试")
        return columnIndex


# 获取excel数据表所有记录
# 首行作为列名称,不记录数据
def getAllItemList(srcExcelFilePath=targetFileName, sheetName=''):
    global wb, targetFileName
    getWorkBook(srcExcelFilePath)
    if not sheetName:
        sheetNames = wb.get_sheet_names()
        sheetSize = len(sheetNames)
        sheetName = sheetNames[sheetSize - 1]
    sheet = wb.get_sheet_by_name(sheetName)
    if sheet:
        allItem = []
        columnSize = len(sheet[1])
        rowSize = len(sheet['A'])
        columnNames = [sheet.cell(row=1, column=x).value for x in range(1, columnSize + 1)]
        for x in range(2, rowSize + 1):
            item = {}
            item['row_index'] = x
            for y in range(1, columnSize):
                item[columnNames[y]] = sheet.cell(row=x, column=y + 1).value
            allItem.append(item)
        return allItem
    else:
        print("指定的sheet(%s)不存在,appendInfo失败,请重试" % sheetName)


def close():
    global wb, targetFileName
    if wb:
        wb.save(targetFileName)
        wb.close()


# appendInfo('', 2, "hello2", "minSdkVersion")
# appendInfo('', 2, "hello3", "targetSdkVersion")
# close()
print("allSize = ", len(getAllItemList()))
# initWorkBook(*[x for x in range(1, 10)])
# appendItem(*["hello", "hello", "hello", "hello", "hello"])
